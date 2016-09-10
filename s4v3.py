import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from s4v2 import *

print("\n s4v3: ")

#create function to save excel workbook
def save_spreadsheet(filename, lst):
	wb = Workbook()
	ws = wb.active

	row_index = 1
	for rows in lst:
		col_index = 1 
		for field in rows:
			col_letter = get_column_letter(col_index)
			ws.cell('{}{}'.format(col_letter, row_index)).value = field
			col_index += 1
		row_index += 1

	wb.save(filename)

kiton_ties = filter_col_by_string(data_from_csv, "brandName", "Kiton")
save_spreadsheet("_data/s4_kiton.xlsx", kiton_ties)

print("\n writing _data/s4_kiton.xlsx finished")